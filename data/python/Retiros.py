import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import pyperclip
import os
import sys
import re

# Configuración inicial (archivos y datos de entrada)
topisimoname = 'data/excel/topisimo.xlsx'
bominame = 'data/excel/bomi.xlsx'
excelretoriginal = 'data/excel/SolicitudRetiroTeva.xlsx'
ruta_escritorio = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')


def update_progress(progress):
    with open(progreso, "w") as f:
        f.write(str(progress))


textoenviado = sys.argv[1]
progreso = sys.argv[2]
update_progress(10)

# --- Captura de datos ---
datos=[]
formulario=[
("Nombre establecimiento"),
("RUT establecimiento"),
("Dirección para retirar los productos"),
("Nombre del contacto"),
("E-mail del contacto"),
("Fono de contacto"),
("Folio Documento Tributario (guía o factura)"),
("Ventana horaria"),
("Estado del producto")]





ingresopaste=textoenviado



lineas=ingresopaste.split('\n')
for linea in lineas:
    linea=linea.strip()
    if linea=='':
        pass
    else:
        datos.append(linea)

datos.insert(0,"CANJE")


datos[7]=re.sub(r'\D', '', datos[7])  # Reemplaza todo lo que NO sea dígito


update_progress(15)

nombrearchivo = f"Solicitud retiro Teva - {datos[1]} - {datos[7]}"

ruta_copia = f"{ ruta_escritorio+'/'+nombrearchivo}.xlsx"

# --- Función de búsqueda con verificación ---
def buscar_valor_con_verificacion(valor_buscar, archivo, sheet, columnas_busqueda, columna_resultado, columna_verificacion, valor_verificacion):
    """
    Busca un valor en un archivo Excel y verifica que otra columna coincida con un valor dado.
    """
    try:
        libro = openpyxl.load_workbook(archivo)
        hoja = libro[sheet]
        
        for fila in hoja.iter_rows(values_only=True):
            # Buscar en columnas_busqueda
            for col in columnas_busqueda:
                indice_col = openpyxl.utils.column_index_from_string(col) - 1
                if str(fila[indice_col]) == str(valor_buscar):
                    # Verificar columna adicional
                    indice_verificacion = openpyxl.utils.column_index_from_string(columna_verificacion) - 1
                    if str(fila[indice_verificacion]) == str(valor_verificacion):
                        indice_resultado = openpyxl.utils.column_index_from_string(columna_resultado) - 1
                        return fila[indice_resultado]
        return None
    except Exception as e:
        print(f"Error en {archivo}: {e}")
        return None

# --- Variables clave ---
valor_a_buscar = int(datos[7])
valor_verificacion = datos[2].replace('.','') # Ejemplo: Número de documento del cliente

update_progress(25)
# Columnas de búsqueda y verificación (personaliza según tu Excel)
config_bomi = {
    "columnas_busqueda": ["N", "D"],
    "columna_verificacion": "A",  # Columna para verificar (ej: documento)
    "resultados": {"prod": "G", "sku": "H", "lote": "I"}
}

config_topisi = {
    "columnas_busqueda": ["C", "D"],
    "columna_verificacion": "Q",  # Columna para verificar en Topisimo
    "resultados": {"prod": "AI", "sku": "N", "lote": "P"}
}


# --- Búsqueda principal ---
resultados = {"prod": None, "sku": None, "lote": None}

# 1. Buscar en BOMI
for key, col in config_bomi["resultados"].items():
    resultados[key] = buscar_valor_con_verificacion(
        valor_a_buscar, bominame,"Sheet1",
        config_bomi["columnas_busqueda"],
        col,
        config_bomi["columna_verificacion"],
        valor_verificacion
    )


# 2. Buscar en TOPISI si no se encontró en BOMI
if resultados["prod"] is None:
    for key, col in config_topisi["resultados"].items():
        resultados[key] = buscar_valor_con_verificacion(
            valor_a_buscar, topisimoname,"CL Sales",
            config_topisi["columnas_busqueda"],
            col,
            config_topisi["columna_verificacion"],
            valor_verificacion
        )
update_progress(50)
# --- Resultados ---
if resultados["prod"] is not None:
    print(f"Producto: {resultados['prod']}")
    print(f"SKU: {resultados['sku']}")
    print(f"Lote: {resultados['lote']}")
else:
    print("No se encontraron coincidencias válidas.")

# --- Modificar el archivo de salida ---
libro = openpyxl.load_workbook(excelretoriginal)
hoja = libro['Hoja1']


# Datos fijos
hoja['D14'] = datos[1]
hoja['D15'] = datos[3]
hoja['D16'] = datos[4]
hoja['D17'] = datos[5]
hoja['D18'] = datos[6]
hoja['J14'] = datos[7]
hoja['J16'] = datos[8]
hoja['J17'] = datos[9]
hoja['J18'] = datos[0]

# Datos buscados
hoja['J22'] = resultados["prod"]
hoja['J23'] = resultados["sku"]
hoja['J24'] = resultados["lote"]

libro.save(ruta_copia)
print(f"¡Archivo guardado en {ruta_copia}!")
update_progress(100)