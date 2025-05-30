import requests
from bs4 import BeautifulSoup
from ftfy import fix_text
import json
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
import subprocess
import sys
import os
import datetime,time

ruta_escritorio = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
rutiniciosesion = sys.argv[4]
contraseñainiciosesion = sys.argv[5]

#URL del Login (reemplázala con la real)
URL_LOGIN = "https://aplicacionesweb.cenabast.cl/new-sistema/OIRS2/asp/login.asp"
URL_INICIO = "https://aplicacionesweb.cenabast.cl/new-sistema/OIRS2/inicio.asp"  # Después del login
# URL de la API AJAX 
API_LISTA_RECLAMOS = "https://aplicacionesweb.cenabast.cl/new-sistema/oirs2/modulos/BandejaAsignados/SolicitudesClientes.asp?page=undefined"

#Nombre de archivos topisimo y bomi
topisimoname='data/excel/topisimo.xlsx'
bominame='data/excel/bomi.xlsx'
'''
# Genera un enlace directo (usa "Embed" y copia la URL src)
onedrive_url = 'https://...enlace.../content'
df = pd.read_excel(onedrive_url)
'''

#Datos del usuario
CREDENCIALES = {
    "usuario": {rutiniciosesion},
    "password": {contraseñainiciosesion},
}

#Headers opcionales
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/133.0.0.0 Safari/537.36",
    "Referer": "https://aplicacionesweb.cenabast.cl/new-sistema/OIRS2/index.asp",
    "Content-Type": "application/x-www-form-urlencoded"
}

#fechahoy
hoy=datetime.datetime.today().strftime("%d-%m-%Y")

#Ruta ejecutable
python_path = sys.executable

def iniciar_sesion():
    """Inicia sesión en el sistema y devuelve una sesión autenticada."""
    session = requests.Session()  #Mantiene la sesión activa

    #Paso 1: Enviar POST al login
    respuesta = session.post(URL_LOGIN, data=CREDENCIALES, headers=HEADERS, allow_redirects=False)

    #Paso 2: Verificar si el login fue exitoso (Código 302 = Redirección)
    if respuesta.status_code == 302 and "location" in respuesta.headers:
        print("Login exitoso. Redirigiendo a inicio.asp...")
        
        #Paso 3: Seguir la redirección para completar la autenticación
        session.get(URL_INICIO, headers=HEADERS)

        return session
    else:
        print("Error al iniciar sesión:", respuesta.status_code)
        print("Respuesta del servidor:", respuesta.text)
        return None

def obtener_lista_reclamos(session):
    """Obtiene la lista de reclamos autenticado."""
    respuesta = session.get(API_LISTA_RECLAMOS, headers=HEADERS)

    print("Código de respuesta HTTP:", respuesta.status_code)
    #print("Contenido de la respuesta:", respuesta.text)  # Verificamos que la respuesta sea JSON

    if respuesta.status_code != 200:
        print("Error al obtener la lista de reclamos:", respuesta.status_code)
        return []

    try:
        soup = BeautifulSoup(respuesta.text, "html.parser")
        datos = []
        for elemento in soup.find_all("tr"):
            texto = elemento.contents[1].get_text(strip=True)
            texto=texto[0:6]
            datos.append(texto)
        del datos[0]
        return datos

        #return respuesta.json()
    except requests.exceptions.JSONDecodeError:
        print("⚠️ Error: La API no devolvió JSON válido.")
        return []

def obtener_detalle_reclamo(session, num_solicitud):
    alldatos=[]
    count=1
    count2=0
    print("======= Obteniendo Datos Solicitados =======")
    for num in num_solicitud:
        url_detalles = f"https://aplicacionesweb.cenabast.cl/new-sistema/OIRS2/Modulos/BandejaAsignados/SolicitudesClientes_2.asp?numSolicitud={num}"
        
        # Enviar la solicitud con la sesión activa
        respuesta = session.get(url_detalles)

        if respuesta.status_code == 200:

            print(f"✅ Detalles del reclamo {num},{count} de {len(num_solicitud)} obtenidos correctamente.")
            
            porcent=len(num_solicitud)
            porcent=60/porcent
            
            
            update_progress(int((porcent*(count-1))+10))
          
         


            soup = BeautifulSoup(respuesta.text, "html.parser")
            datos = {}
            
            elemento = soup.find("table")
            datos['Solicitud']=num
            datos['Rut Solicitante']=elemento.contents[1].contents[3].get_text(strip=True)[2:]
            datos['Solicitante']=elemento.contents[1].contents[7].get_text(strip=True)[2:]
            datos['Destinatario']=elemento.contents[3].contents[7].get_text(strip=True)[2:]
            datos['Tipo Solicitud']=elemento.contents[5].contents[3].get_text(strip=True)[2:]
            datos['Usuario']=elemento.contents[7].contents[3].get_text(strip=True)[2:]
            datos['Contacto']=elemento.contents[7].contents[7].get_text(strip=True)[2:]
            datos['DV']=elemento.contents[9].contents[3].get_text(strip=True)[2:]
            datos['GD']='-'
            datos['FA / NC']='-'
            datos['Repeticiones DV']=''
            datos['Fecha']=elemento.contents[11].contents[3].get_text(strip=True)[2:12]
            datos['Proveedor']=elemento.contents[13].contents[3].get_text(strip=True)[2:]
            datos['Producto']=elemento.contents[13].contents[7].get_text(strip=True)[2:]
            datos['CategorIA']='-'
            datos['Comentario']=elemento.contents[15].contents[3].get_text(strip=True)[2:]
            
            for dato in datos:
                if datos[dato]=='':
                    datos[dato]='-'
                texto_corregido=fix_text(datos[dato])
                datos[dato]=texto_corregido
                if dato =="Comentario":
                    comentario=datos[dato]
                    comentario=normalizar_comentario_mejorado(comentario)
                    datos[dato]=comentario

            alldatos.append(datos)
        else:
                print(f"⚠️ Error al obtener el reclamo {num_solicitud}: {respuesta.status_code}")
                return None
        count+=1
        
    #alldatos=json.dumps(alldatos,ensure_ascii=False, indent=4)
    return alldatos

def normalizar_comentario_mejorado(comentario):
    # Convertir todo el texto a minúsculas
    comentario = comentario.lower()

    # Lista de palabras que deben mantenerse en mayúsculas
    excepciones = ["n°", "lab", "chile", "sa", "pharmatrade"]

    # Dividir el texto en oraciones usando expresiones regulares
    oraciones = re.split(r'(?<=[.!?])\s+', comentario)

    # Capitalizar la primera letra de cada oración y manejar excepciones
    oraciones_normalizadas = []
    for oracion in oraciones:
        palabras = oracion.split()
        palabras_normalizadas = []
        for palabra in palabras:
            if palabra in excepciones:
                palabras_normalizadas.append(palabra.upper())
            else:
                palabras_normalizadas.append(palabra)
        oracion_normalizada = ' '.join(palabras_normalizadas)
        oracion_normalizada = oracion_normalizada.capitalize()
        oraciones_normalizadas.append(oracion_normalizada)

    # Unir las oraciones nuevamente
    comentario_normalizado = ' '.join(oraciones_normalizadas)

    return comentario_normalizado

def obtener_info_topisimo(dv, df_topisimo, dv_counts_topisimo):
    """
    Obtiene la información de GD, FA/NC y el contador de repeticiones de DV
    para un DV específico en el archivo "Topisimo".
    """
    # Normalizar el DV (eliminar espacios y convertir a cadena)
    dv = str(dv).strip()

    # Buscar el DV en la columna "DV 123" (columna A)
    coincidencias = df_topisimo[df_topisimo["DV 123"].astype(str).str.strip() == dv]
    
    if not coincidencias.empty:
        # Tomar la primera fila que coincide (incluso si hay duplicados)
        primera_fila = coincidencias.iloc[0]
        
        # Extraer los valores de las columnas "GD" (C) y "FA / NC" (D)
        gd = primera_fila.get("GD", "")  # Columna C
        fa_nc = primera_fila.get("FA / NC", "")  # Columna D
        
        # Obtener el número de repeticiones del DV
        repeticiones = dv_counts_topisimo.get(dv, 0)
        
        return gd, fa_nc, repeticiones
    return "", "", 0  # Si no se encuentra el DV, devolver valores vacíos

def obtener_info_bomi(dv, df_bomi, dv_counts_bomi):
    """
    Obtiene la información de GD y FA para un DV específico en el archivo "Bomi".
    """
    # Normalizar el DV (eliminar espacios y convertir a cadena)
    dv = str(dv).strip()

    # Buscar el DV en la columna "DV" (columna P)
    coincidencias = df_bomi[df_bomi["DV"].astype(str).str.strip() == dv]
    
    if not coincidencias.empty:
        # Tomar la primera fila que coincide (incluso si hay duplicados)
        primera_fila = coincidencias.iloc[0]
        
        # Extraer los valores de las columnas "GD" y "FA"
        gd = primera_fila.get("GD", "")  # Columna GD
        fa = primera_fila.get("FA", "")  # Columna FA
        
        # Obtener el número de repeticiones del DV
        repeticiones = dv_counts_bomi.get(dv, 0)
        
        return gd, fa, repeticiones
    return "", "", 0  # Si no se encuentra el DV, devolver valores vacíos

def cruce(datos_reclamos):
    """
    Realiza el cruce de datos entre los reclamos y los archivos "Topisimo" y "Bomi".
    """
    print("===== Generando Archivo Excel =====")
    # Convertir los datos de reclamos a un DataFrame
    df_reclamos = pd.DataFrame(datos_reclamos)

    # Leer el archivo "Topisimo"
    df_topisimo = pd.read_excel(topisimoname,sheet_name='CL Sales')

    # Leer el archivo "Bomi"
    df_bomi = pd.read_excel(bominame)

    # Normalizar nombres de columnas en "Topisimo" y "Bomi" (eliminar espacios adicionales)
    df_topisimo.columns = df_topisimo.columns.str.strip()
    df_bomi.columns = df_bomi.columns.str.strip()

    # Verificar que las columnas necesarias existan en "Topisimo"
    columnas_requeridas_topisimo = ["DV 123", "GD", "FA / NC"]
    for columna in columnas_requeridas_topisimo:
        if columna not in df_topisimo.columns:
            raise ValueError(f"⚠️ Error: La columna '{columna}' no existe en el archivo 'Topisimo'.")

    # Verificar que las columnas necesarias existan en "Bomi"
    columnas_requeridas_bomi = ["DV", "GD", "FA"]
    for columna in columnas_requeridas_bomi:
        if columna not in df_bomi.columns:
            raise ValueError(f"⚠️ Error: La columna '{columna}' no existe en el archivo 'Bomi'.")

    # Contar las repeticiones de cada DV en "Topisimo" y "Bomi"
    dv_counts_topisimo = df_topisimo["DV 123"].astype(str).str.strip().value_counts()
    dv_counts_bomi = df_bomi["DV"].astype(str).str.strip().value_counts()

    # Agregar columnas vacías para GD, FA/NC y contador de repeticiones
    df_reclamos["GD"] = "-"
    df_reclamos["FA / NC"] = "-"
    df_reclamos["Repeticiones DV"] = "-"
    df_reclamos["Fuente"] = "-"  # Columna para indicar de dónde se obtuvieron los datos

    # Realizar el cruce de datos
    for index, row in df_reclamos.iterrows():
        dv = row["DV"]
        if dv:  # Solo procesar si el DV no está vacío
            # Buscar primero en "Topisimo"
            gd, fa_nc, repeticiones = obtener_info_topisimo(dv, df_topisimo, dv_counts_topisimo)
            if gd or fa_nc:  # Si se encontró información en "Topisimo"
                df_reclamos.at[index, "GD"] = gd
                df_reclamos.at[index, "FA / NC"] = fa_nc
                df_reclamos.at[index, "Repeticiones DV"] = repeticiones
                df_reclamos.at[index, "Fuente"] = "Topisimo"
            else:
                # Si no se encontró en "Topisimo", buscar en "Bomi"
                gd, fa, repeticiones = obtener_info_bomi(dv, df_bomi, dv_counts_bomi)
                if gd or fa:  # Si se encontró información en "Bomi"
                    df_reclamos.at[index, "GD"] = gd
                    df_reclamos.at[index, "FA / NC"] = fa
                    df_reclamos.at[index, "Repeticiones DV"] = repeticiones
                    df_reclamos.at[index, "Fuente"] = "Bomi"


    # Aplicar formato con la función excelbonito
    excelbonito(df_reclamos)

    print("✅ Excel generado y formateado correctamente.")

def excelbonito(df):
    """
    Aplica formato a un DataFrame y lo guarda en un archivo Excel.
    """

    # Guardamos a un archivo Excel
    archivo_excel = ruta_escritorio+"/reclamos.xlsx"
    archivo_excel=guardar_excel_con_rotacion()
    df.to_excel(archivo_excel, index=False, engine="openpyxl")

    # 📌 Estilizamos el archivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Reclamos"
    for r_idx, row in enumerate(df.columns.tolist(), 1):
        ws.cell(row=1, column=r_idx, value=row)

    # Agregar datos
    for r_idx, row in enumerate(df.itertuples(index=False), 2):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # 📌 Estilos para encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # 📌 Ajuste de ancho automático de columnas
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # 📌 Aplicar formato de tabla
    tabla = Table(displayName="TablaReclamos", ref=ws.dimensions)
    estilo_tabla = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tabla.tableStyleInfo = estilo_tabla
    ws.add_table(tabla)

    # 📌 Guardamos el archivo con formato aplicado
    wb.save(archivo_excel)
    print(f"✅✅ Archivo '{archivo_excel}' creado con formato de tabla.✅✅")
    
def guardar_json(datos, archivo='data/json/inforeclamos.json'):
    with open(archivo, 'w', encoding='utf-8') as f:
        json.dump(datos, f, ensure_ascii=False, indent=4)
    print(f"✅ Datos guardados en {archivo}")

def update_progress(progress):
    with open(progreso, "w") as f:
        f.write(str(progress))


def guardar_excel_con_rotacion( nombre_base="reclamos"):
    """
    Guarda los datos en un Excel con rotación diaria:
    """


    # Nombre de archivos
    archivo_actual = f"{ ruta_escritorio+'/'+nombre_base}.xlsx"
    
    # Si existe el archivo actual, renombrarlo a histórico
    countexcel=1
    while True:
        if os.path.exists(archivo_actual):
             archivo_actual = f"{ ruta_escritorio+'/'+nombre_base}({countexcel}).xlsx"
             countexcel+=1
        else:
            break
    return(archivo_actual)

 

if __name__ == "__main__":
    try:
        textoenviado = sys.argv[1]
        progreso = sys.argv[2]
        update_progress(10)
        session = iniciar_sesion()

        if session:
            if textoenviado!='':
                reclamos=[]
                lineas=textoenviado.split('\n')
                for linea in lineas:
                    linea=linea.strip()
                    if linea=='':
                        pass
                    else:
                        reclamos.append(linea)
            else:
                reclamos = obtener_lista_reclamos(session)

            # Para pruebas usar "testreclamos", si no es prueba, usar "reclamos", si hay varios reclamos, usar "reclamosaleat" y pegar los datos en json reclamosaleat.json
            datos_reclamos = obtener_detalle_reclamo(session, reclamos)

            guardar_json(datos_reclamos)

            print('======= Inicio Proceso IA =======')
            update_progress(80)
            subprocess.run([python_path, "data/python/categorizacion.py"])
            update_progress(90)
            print('======= Fin Proceso IA ======')

            #Cargar json con categorIA
            with open('data/json/inforeclamos.json', 'r', encoding='utf-8') as f:
                datos_reclamos = json.load(f)

            # Intenta crear el DataFrame solo si tiene contenido válido
            if isinstance(datos_reclamos, list) and all(isinstance(i, dict) for i in datos_reclamos):
            
                cruce(datos_reclamos)

            else:
                print("❌ Error: datos_reclamos no tiene el formato correcto.")
            
            update_progress(100)
    except Exception as e:
        import traceback
        print("❌ Error")
        traceback.print_exc()
        input("Presiona Enter para salir...")
        
        






